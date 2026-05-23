"""
Build the AI & ML taxonomy v3 migration SQL from the v7-final Excel.

Drops Excel L5 rows (Option C). Each node's DB level is computed as
  db_level = parent.db_level + 1
(where the root 'ai-ml' L1 is DB level 1). This naturally absorbs Excel
"level skips" (a few L3s sit directly under an L1, a few L4s sit directly
under an L2) without inventing placeholder categories.

For the normal full chain (Excel L1→L2→L3→L4), this gives DB L2→L3→L4→L5
exactly as planned.

Produces:
  database/migration-aiml-taxonomy-v3.sql     — sections A-E (Section F is appended later by the listing-remap step)
  exports/aiml-taxonomy-v3.json               — parsed tree for the listing remap step

Slugs:
  - slugify(name) → lowercase, [^a-z0-9]+ → '-', trim hyphens, truncate to 80
  - Collisions WITHIN the new set      → append -2, -3, ...
  - Collisions vs LIVE non-AI&ML slugs → append -ai-ml (then -2, -3 if needed)

Run:
  python scripts/build-aiml-v3-migration.py
"""
import json
import re
import sys
import io
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path(r"C:\Users\AADIL PARMAR\Downloads\AI ML - hierarchy v7 Clau  final.xlsx")
SQL_OUT = ROOT / "database" / "migration-aiml-taxonomy-v3.sql"
JSON_OUT = ROOT / "exports" / "aiml-taxonomy-v3.json"
NON_AIML_SLUGS_FILE = ROOT / "exports" / "non-aiml-slugs.json"

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

AIML_COLOR = "#8B5CF6"           # current AI&ML purple
AIML_PARENT_SLUG = "ai-ml"       # live L1 slug
AIML_PARENT_LEVEL = 1


def slugify(name: str) -> str:
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    if len(s) > 80:
        s = s[:80].rstrip("-")
    return s


def sql_escape(s: str) -> str:
    return (s or "").replace("\\", "\\\\").replace("'", "''")


# ─── 1. Load existing non-AI&ML slugs ─────────────────────────────────────────
existing = json.loads(NON_AIML_SLUGS_FILE.read_text(encoding="utf-8"))
EXISTING_SLUGS = {row["slug"] for row in existing["non_aiml_slugs"]}
print(f"Loaded {len(EXISTING_SLUGS)} existing non-AI&ML slugs to avoid colliding with.")


# ─── 2. Read the Excel ────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["Merged Taxonomy"]
print(f"Sheet: Merged Taxonomy  rows={ws.max_row}  cols={ws.max_column}")

LEVELS = ["L1", "L2", "L3", "L4", "L5"]

raw_rows = []
for r in range(2, ws.max_row + 1):
    sr = ws.cell(row=r, column=1).value
    vals = {lvl: ws.cell(row=r, column=2 + i).value for i, lvl in enumerate(LEVELS)}
    prim = ws.cell(row=r, column=7).value
    raw_rows.append((sr, vals, prim))

print(f"Total Excel rows: {len(raw_rows)}")


# ─── 3. Walk the Excel, tracking cursor at each level. Drop L5. ───────────────
# Each kept node remembers its full chain of (excel_level, name) ancestors.
cur = {lvl: None for lvl in LEVELS}
nodes = []
counts = {lvl: 0 for lvl in LEVELS}
counts["DROP_L5"] = 0

for sr, vals, prim in raw_rows:
    own = vals.get(prim)
    if own is None:
        continue
    own = str(own).strip()
    if not own:
        continue
    cur[prim] = own
    idx = LEVELS.index(prim)
    for deeper in LEVELS[idx + 1 :]:
        cur[deeper] = None
    counts[prim] += 1

    if prim == "L5":
        counts["DROP_L5"] += 1
        continue

    ancestor_chain = [(lvl, cur[lvl]) for lvl in LEVELS[:idx] if cur[lvl] is not None]
    nodes.append(
        {
            "excel_level": prim,
            "name": own,
            "ancestors": ancestor_chain,
            "sr": sr,
        }
    )

print("Excel counts:", counts)
print(f"Nodes kept (L1-L4 only): {len(nodes)}")


# ─── 4. Generate slugs with collision detection ───────────────────────────────
used_in_new = set()


def pick_slug(base: str) -> str:
    if not base:
        base = "category"
    if base not in used_in_new and base not in EXISTING_SLUGS:
        used_in_new.add(base)
        return base
    if base in EXISTING_SLUGS:
        cand = f"{base}-ai-ml"
        if len(cand) > 80:
            cand = cand[:80].rstrip("-")
        if cand not in used_in_new and cand not in EXISTING_SLUGS:
            used_in_new.add(cand)
            return cand
    for n in range(2, 1000):
        suffix = f"-{n}"
        cand = base if len(base) + len(suffix) <= 80 else base[: 80 - len(suffix)].rstrip("-")
        cand = f"{cand}{suffix}"
        if cand not in used_in_new and cand not in EXISTING_SLUGS:
            used_in_new.add(cand)
            return cand
    raise RuntimeError(f"could not allocate unique slug for {base!r}")


# Index: (own_excel_level, own_name) + tuple(ancestor names) → final_slug + db_level
slug_index = {}
emitted = []
sort_counter = {}  # key: parent_path tuple → next sort_order

for node in nodes:
    ancestors = node["ancestors"]
    own_lvl = node["excel_level"]
    own_name = node["name"]

    # Resolve parent: either an ancestor (if any) or the AI&ML L1 root.
    if not ancestors:
        parent_slug = AIML_PARENT_SLUG
        parent_db_level = AIML_PARENT_LEVEL  # 1
    else:
        parent_lvl_excel, parent_name = ancestors[-1]
        parent_key = (parent_lvl_excel, parent_name) + tuple(a[1] for a in ancestors[:-1])
        parent_info = slug_index.get(parent_key)
        if not parent_info:
            raise RuntimeError(f"parent_info missing for {node!r}  (parent_key={parent_key})")
        parent_slug = parent_info["slug"]
        parent_db_level = parent_info["db_level"]

    db_level = parent_db_level + 1
    if db_level > 5:
        # Should not happen with Option C (no Excel L5), but be safe.
        raise RuntimeError(f"db_level {db_level} > 5 for node {node!r}")

    base = slugify(own_name)
    final = pick_slug(base)

    parent_path_tuple = tuple(a[1] for a in ancestors)
    sort_counter[parent_path_tuple] = sort_counter.get(parent_path_tuple, 0) + 10

    rec = {
        "excel_level": own_lvl,
        "db_level": db_level,
        "name": own_name,
        "slug": final,
        "ancestors": [{"level": l, "name": n} for l, n in ancestors],
        "parent_slug": parent_slug,
        "parent_db_level": parent_db_level,
        "sort_order": sort_counter[parent_path_tuple],
        "sr": node["sr"],
    }
    emitted.append(rec)
    slug_index[(own_lvl, own_name) + parent_path_tuple] = {"slug": final, "db_level": db_level}

print(f"Emitted {len(emitted)} nodes with final slugs.")
by_db_level = {f"L{lvl}": sum(1 for n in emitted if n["db_level"] == lvl) for lvl in (2, 3, 4, 5)}
print("By DB level:", by_db_level)


# ─── 5. JSON output for the listing-remap step ────────────────────────────────
JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
JSON_OUT.write_text(
    json.dumps({"generated_from": EXCEL.name, "by_db_level": by_db_level, "nodes": emitted}, indent=2, ensure_ascii=False),
    encoding="utf-8",
)
print(f"Wrote {JSON_OUT}  ({len(emitted)} nodes)")


# ─── 6. Emit the SQL migration ────────────────────────────────────────────────
lines = []
lines.append("-- ============================================================")
lines.append("-- InfoWebWorld — AI & ML Taxonomy v3 Migration")
lines.append("-- Rebuilds the AI & ML sector with 1,392 hierarchical categories")
lines.append("-- across 4 nested levels (DB L2..L5 under existing 'ai-ml' L1).")
lines.append("-- Drops the flat listing_types rows for AI&ML.")
lines.append("--")
lines.append("-- Source: AI ML - hierarchy v7 Clau  final.xlsx")
lines.append("-- Run each section IN ORDER in phpMyAdmin.")
lines.append("-- ============================================================")
lines.append("")
lines.append("-- ═══ Section A: Safety ═════════════════════════════════════════")
lines.append("SET FOREIGN_KEY_CHECKS = 0;")
lines.append("")

lines.append("-- ═══ Section B: Disconnect existing AI&ML submissions ═════════")
lines.append("-- Listings are PRESERVED. We null out category_id + listing_type_id")
lines.append("-- here, then re-attach them in Section F (appended at the bottom).")
lines.append("UPDATE submissions")
lines.append("   SET category_id = NULL, listing_type_id = NULL")
lines.append(" WHERE category_id IN (")
lines.append("   SELECT id FROM (")
lines.append("     SELECT c.id FROM categories c")
lines.append("      LEFT JOIN categories p   ON p.id = c.parent_id")
lines.append("      LEFT JOIN categories gp  ON gp.id = p.parent_id")
lines.append("      LEFT JOIN categories ggp ON ggp.id = gp.parent_id")
lines.append("      WHERE c.parent_id   = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1)")
lines.append("         OR p.parent_id  = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1)")
lines.append("         OR gp.parent_id = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1)")
lines.append("         OR ggp.parent_id = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1)")
lines.append("   ) AS aiml_ids")
lines.append(" );")
lines.append("")

lines.append("-- ═══ Section C: Delete old AI&ML dependents + categories ══════")
lines.append("-- Order: SEO content → listing_types → child cats → parent cats.")
lines.append("")
lines.append("-- C.1: delete category_seo_content for old AI&ML categories")
lines.append("DELETE sc FROM category_seo_content sc")
lines.append("  JOIN categories c ON c.id = sc.category_id")
lines.append("  LEFT JOIN categories p  ON p.id = c.parent_id")
lines.append("  LEFT JOIN categories gp ON gp.id = p.parent_id")
lines.append(" WHERE c.parent_id  = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1)")
lines.append("    OR p.parent_id  = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1)")
lines.append("    OR gp.parent_id = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1);")
lines.append("")
lines.append("-- C.2: delete listing_types tied to old AI&ML categories")
lines.append("DELETE lt FROM listing_types lt")
lines.append("  JOIN categories c ON c.id = lt.category_id")
lines.append("  LEFT JOIN categories p ON p.id = c.parent_id")
lines.append(" WHERE c.parent_id = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1)")
lines.append("    OR p.parent_id = (SELECT id FROM categories WHERE slug='ai-ml' AND level=1);")
lines.append("")
lines.append("-- C.3: delete L3 categories under AI&ML (JOIN-based to dodge MySQL #1093)")
lines.append("DELETE c FROM categories c")
lines.append("  JOIN categories p  ON p.id  = c.parent_id")
lines.append("  JOIN categories gp ON gp.id = p.parent_id")
lines.append(" WHERE c.level = 3 AND gp.slug = 'ai-ml' AND gp.level = 1;")
lines.append("")
lines.append("-- C.4: delete L2 categories under AI&ML — root 'ai-ml' L1 stays. JOIN-based.")
lines.append("DELETE c FROM categories c")
lines.append("  JOIN categories p ON p.id = c.parent_id")
lines.append(" WHERE c.level = 2 AND p.slug = 'ai-ml' AND p.level = 1;")
lines.append("")

for db_lvl in (2, 3, 4, 5):
    bucket = [n for n in emitted if n["db_level"] == db_lvl]
    lines.append(f"-- ═══ Section D.{db_lvl - 1}: Insert {len(bucket)} new L{db_lvl} categories ═══")
    for n in bucket:
        name_sql = sql_escape(n["name"])
        slug_sql = sql_escape(n["slug"])
        parent_slug_sql = sql_escape(n["parent_slug"])
        lines.append(
            "INSERT INTO categories (name, slug, level, parent_id, color, is_active, is_launched, is_navigation, sort_order)"
        )
        lines.append(
            f"SELECT '{name_sql}', '{slug_sql}', {db_lvl}, id, '{AIML_COLOR}', 1, 1, 1, {n['sort_order']}"
        )
        lines.append(
            f"  FROM categories WHERE slug = '{parent_slug_sql}' AND level = {n['parent_db_level']} LIMIT 1;"
        )
    lines.append("")

lines.append("-- ═══ Section E: Re-enable FKs ═════════════════════════════════")
lines.append("SET FOREIGN_KEY_CHECKS = 1;")
lines.append("")
lines.append("-- ═══ Section F (appended by listing-remap step) ═══════════════")
lines.append("-- See bottom of file for UPDATE statements re-attaching the live")
lines.append("-- AI&ML submissions to their best-fit new categories.")
lines.append("")

SQL_OUT.parent.mkdir(parents=True, exist_ok=True)
SQL_OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {SQL_OUT}  ({SQL_OUT.stat().st_size:,} bytes)")
