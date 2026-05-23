"""Export AI & ML taxonomy from migration SQL to CSV + styled Excel.

Levels:
  L1 = root (Artificial Intelligence & ML)
  L2 = 43 sub-categories       (migration-aiml-taxonomy-v2.sql)
  L3 = 1,295 sub-sub-categories (migration-aiml-taxonomy-v2.sql)
  L4 = 2,050 listing types     (migration-remap-aiml-listing-types.sql)
       — each listing type uses 'Group: Specific' naming, which we split:
         L4 Group  = part before the colon (acts like a facet / mini-L5)
         L4 Type   = full listing-type name
"""
import re
import csv
import sys
import io
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parents[1]
SQL_CATS = ROOT / "database" / "migration-aiml-taxonomy-v2.sql"
SQL_LT = ROOT / "database" / "migration-remap-aiml-listing-types.sql"
OUT_CSV = ROOT / "exports" / "aiml-taxonomy-full.csv"
OUT_XLSX = ROOT / "exports" / "aiml-taxonomy-full.xlsx"
OUT_CSV.parent.mkdir(parents=True, exist_ok=True)

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

txt = SQL_CATS.read_text(encoding="utf-8")
lt_txt = SQL_LT.read_text(encoding="utf-8")

# Capture name, slug, level, parent_slug, parent_level + sort_order
pat = re.compile(
    r"SELECT '([^']+)', '([^']+)', (\d+), id, 1, 1, 1, (\d+)\s*"
    r"FROM categories WHERE slug = '([^']+)' AND level = (\d+)",
    re.DOTALL,
)
rows = pat.findall(txt)

children = defaultdict(list)  # parent_slug -> [(sort, name, slug, level)]
for name, slug, lvl, sort_order, pslug, plvl in rows:
    children[pslug].append((int(sort_order), name, slug, int(lvl)))

for k in children:
    children[k].sort(key=lambda x: x[0])

# Parse listing types (L4): each row references an L3 category by slug.
# Format: ((SELECT id FROM categories WHERE slug = 'L3-SLUG' AND level = 3 LIMIT 1),
#          'L4 NAME', 'L4-SLUG', SORT)
lt_pat = re.compile(
    r"slug = '([^']+)' AND level = 3 LIMIT 1\),\s*"
    r"'([^']+)',\s*'([^']+)',\s*(\d+)\)",
    re.DOTALL,
)
listing_types = defaultdict(list)  # L3 slug -> [(sort, name, slug)]
for l3_slug, lt_name, lt_slug, lt_sort in lt_pat.findall(lt_txt):
    listing_types[l3_slug].append((int(lt_sort), lt_name, lt_slug))
for k in listing_types:
    listing_types[k].sort(key=lambda x: x[0])

total_l4 = sum(len(v) for v in listing_types.values())
print(f"Parsed L4 listing types: {total_l4} across {len(listing_types)} L3 parents")

L1_NAME = "Artificial Intelligence & ML"
L1_SLUG = "artificial-intelligence-ml"

def split_lt(name: str) -> tuple[str, str]:
    """'Group: Specific' -> ('Group', 'Specific'). Falls back gracefully."""
    if ":" in name:
        g, s = name.split(":", 1)
        return g.strip(), s.strip()
    return "", name.strip()


flat = []
l2s = children.get(L1_SLUG, [])
for l2_sort, l2_name, l2_slug, _ in l2s:
    l3s = children.get(l2_slug, [])
    if not l3s:
        flat.append({
            "L1": L1_NAME, "L1 Slug": L1_SLUG,
            "L2 #": l2_sort // 10, "L2": l2_name, "L2 Slug": l2_slug,
            "L3 #": "", "L3": "", "L3 Slug": "",
            "L4 #": "", "L4 Group": "", "L4 Type": "", "L4 Slug": "",
        })
        continue
    for l3_sort, l3_name, l3_slug, _ in l3s:
        l4s = listing_types.get(l3_slug, [])
        if not l4s:
            flat.append({
                "L1": L1_NAME, "L1 Slug": L1_SLUG,
                "L2 #": l2_sort // 10, "L2": l2_name, "L2 Slug": l2_slug,
                "L3 #": l3_sort // 10, "L3": l3_name, "L3 Slug": l3_slug,
                "L4 #": "", "L4 Group": "", "L4 Type": "", "L4 Slug": "",
            })
            continue
        for l4_sort, l4_name, l4_slug in l4s:
            grp, spec = split_lt(l4_name)
            flat.append({
                "L1": L1_NAME, "L1 Slug": L1_SLUG,
                "L2 #": l2_sort // 10, "L2": l2_name, "L2 Slug": l2_slug,
                "L3 #": l3_sort // 10, "L3": l3_name, "L3 Slug": l3_slug,
                "L4 #": l4_sort // 10, "L4 Group": grp, "L4 Type": spec, "L4 Slug": l4_slug,
            })

# Write CSV (UTF-8 with BOM so Excel opens it cleanly)
with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(flat[0].keys()))
    w.writeheader()
    w.writerows(flat)
print(f"CSV: {OUT_CSV}  ({len(flat)} rows)")

# Styled Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# Sheet 1: Flat table
ws = wb.active
ws.title = "AI-ML Taxonomy"

headers = list(flat[0].keys())
ws.append(headers)
for row in flat:
    ws.append([row[h] for h in headers])

# Header style
header_fill = PatternFill("solid", fgColor="1F2937")
header_font = Font(name="Inter", bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=False)
thin = Side(border_style="thin", color="E5E7EB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = border

ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

# Column widths
widths = {
    "L1": 30, "L1 Slug": 30,
    "L2 #": 6, "L2": 44, "L2 Slug": 42,
    "L3 #": 6, "L3": 50, "L3 Slug": 50,
    "L4 #": 6, "L4 Group": 28, "L4 Type": 42, "L4 Slug": 52,
}
for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 20)

# Alternating row colors by L2 group (cleaner than per-row striping)
group_colors = ["FFFFFF", "F3F4F6"]
prev_l2 = None
group_idx = -1
l2_font_bold = Font(name="Inter", bold=True, color="111827", size=10.5)
l3_font = Font(name="Inter", color="374151", size=10.5)
mono_font = Font(name="Consolas", color="6B7280", size=9.5)
num_font = Font(name="Inter", color="6B7280", size=10)

for r in range(2, ws.max_row + 1):
    l2_val = ws.cell(row=r, column=4).value
    if l2_val != prev_l2:
        group_idx += 1
        prev_l2 = l2_val
    color = group_colors[group_idx % 2]
    fill = PatternFill("solid", fgColor=color)
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col_idx)
        c.fill = fill
        c.alignment = left if h not in ("L2 #", "L3 #", "L4 #") else center
        c.border = border
        if h in ("L1 Slug", "L2 Slug", "L3 Slug", "L4 Slug"):
            c.font = mono_font
        elif h == "L2":
            c.font = l2_font_bold
        elif h in ("L2 #", "L3 #", "L4 #"):
            c.font = num_font
        else:
            c.font = l3_font

# Sheet 2: L2 summary (with L3 + L4 counts)
ws2 = wb.create_sheet("L2 Summary")
ws2.append(["#", "L2 Category", "Slug", "L3 Count", "L4 Count"])
for col_idx in range(1, 6):
    c = ws2.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = border
ws2.row_dimensions[1].height = 28
ws2.freeze_panes = "A2"

for i, (l2_sort, l2_name, l2_slug, _) in enumerate(l2s, 1):
    l3_list = children.get(l2_slug, [])
    l3_count = len(l3_list)
    l4_count = sum(len(listing_types.get(l3_slug, [])) for _, _, l3_slug, _ in l3_list)
    ws2.append([i, l2_name, l2_slug, l3_count, l4_count])

for r in range(2, ws2.max_row + 1):
    color = group_colors[(r - 2) % 2]
    fill = PatternFill("solid", fgColor=color)
    for col_idx in range(1, 6):
        c = ws2.cell(row=r, column=col_idx)
        c.fill = fill
        c.border = border
        c.alignment = center if col_idx in (1, 4, 5) else left
        c.font = mono_font if col_idx == 3 else l3_font

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 48
ws2.column_dimensions["C"].width = 46
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12

# Totals row
total_row = ws2.max_row + 1
total_l3 = sum(len(children.get(s, [])) for _, _, s, _ in l2s)
total_l4_all = sum(len(listing_types.get(l3_slug, []))
                   for _, _, l2_slug, _ in l2s
                   for _, _, l3_slug, _ in children.get(l2_slug, []))
dark = PatternFill("solid", fgColor="1F2937")
white_bold = Font(name="Inter", bold=True, color="FFFFFF")
for col_idx in range(1, 6):
    ws2.cell(row=total_row, column=col_idx).fill = dark
ws2.cell(row=total_row, column=2, value="TOTAL").font = white_bold
ws2.cell(row=total_row, column=2).alignment = left
ws2.cell(row=total_row, column=4, value=total_l3).font = white_bold
ws2.cell(row=total_row, column=4).alignment = center
ws2.cell(row=total_row, column=5, value=total_l4_all).font = white_bold
ws2.cell(row=total_row, column=5).alignment = center

wb.save(OUT_XLSX)
print(f"XLSX: {OUT_XLSX}  ({len(flat)} L3 rows, {len(l2s)} L2 groups)")
