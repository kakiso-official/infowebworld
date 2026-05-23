"""
Build the Software & SaaS v2 taxonomy migration SQL.

Excel structure: a single sheet "1. Master Taxonomy" with 3 columns:
   L1 Category | L2 Subcategory | L3 Software Type
Every row has all three filled — no orphan / skip-level rows.

Maps:
  Excel L1 → DB L2 (children of 'software-saas' L1 sector)
  Excel L2 → DB L3
  Excel L3 → DB L4

Produces:
  database/migration-software-saas-taxonomy-v2.sql   — sections A-E (Section F appended later by the listing-remap step)
  exports/software-saas-taxonomy-v2.json             — parsed tree for the listing remap step

Slugs:
  - slugify(name) → lowercase, [^a-z0-9]+ → '-', trim hyphens, truncate to 80
  - Collisions WITHIN the new set            → append -2, -3, ...
  - Collisions vs LIVE non-software-saas slugs → append -saas first (then -2/-3 if still clashes)

Run:
  python scripts/build-software-saas-v2-migration.py
"""
import json
import re
import sys
import io
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path(r"C:\Users\AADIL PARMAR\Downloads\Software-saas-taxonomy-full v1 final.xlsx")
SQL_OUT = ROOT / "database" / "migration-software-saas-taxonomy-v2.sql"
JSON_OUT = ROOT / "exports" / "software-saas-taxonomy-v2.json"
NON_SAAS_SLUGS_FILE = ROOT / "exports" / "non-software-saas-slugs.json"

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SAAS_COLOR = "#3B82F6"             # current Software/SaaS blue
SAAS_PARENT_SLUG = "software-saas"
SAAS_PARENT_LEVEL = 1


def slugify(name: str) -> str:
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    if len(s) > 80:
        s = s[:80].rstrip("-")
    return s


def sql_escape(s: str) -> str:
    return (s or "").replace("\\", "\\\\").replace("'", "''")


# ─── 1. Load existing non-software-saas slugs ────────────────────────────────
existing = json.loads(NON_SAAS_SLUGS_FILE.read_text(encoding="utf-8"))
EXISTING_SLUGS = {row["slug"] for row in existing["non_software_saas_slugs"]}
print(f"Loaded {len(EXISTING_SLUGS)} existing non-software-saas slugs to avoid colliding with.")


# ─── 2. Read the Excel ───────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["1. Master Taxonomy"]
print(f"Sheet: 1. Master Taxonomy  rows={ws.max_row}  cols={ws.max_column}")

raw_rows = []
for r in range(2, ws.max_row + 1):
    l1 = ws.cell(row=r, column=1).value
    l2 = ws.cell(row=r, column=2).value
    l3 = ws.cell(row=r, column=3).value
    if not (l1 and l2 and l3):
        continue
    raw_rows.append((str(l1).strip(), str(l2).strip(), str(l3).strip()))

print(f"Valid Excel rows (all 3 columns filled): {len(raw_rows)}")


# ─── 3. Build hierarchical nodes ─────────────────────────────────────────────
# Walk rows top-to-bottom. Each row contributes:
#   - its L1 (if first time seen)
#   - its L2 under that L1 (if first time seen under that L1)
#   - its L3 under that (L1, L2) (always — every row is an L3)

# We preserve Excel order for sort_order calculation.
seen_l1 = {}             # L1 name → ordered position
seen_l2 = {}             # (L1, L2) → ordered position within L1
seen_l3 = set()          # (L1, L2, L3) — track uniqueness

l1_records = []          # list of L1 names in encounter order
l2_records = []          # list of (l1, l2) in encounter order
l3_records = []          # list of (l1, l2, l3) in encounter order

for l1, l2, l3 in raw_rows:
    if l1 not in seen_l1:
        seen_l1[l1] = len(l1_records)
        l1_records.append(l1)
    if (l1, l2) not in seen_l2:
        seen_l2[(l1, l2)] = len(l2_records)
        l2_records.append((l1, l2))
    triple = (l1, l2, l3)
    if triple not in seen_l3:
        seen_l3.add(triple)
        l3_records.append(triple)

print(f"Unique L1: {len(l1_records)}  L2: {len(l2_records)}  L3: {len(l3_records)}")


# ─── 4. Generate slugs with collision detection ──────────────────────────────
used_in_new = set()


def pick_slug(base: str) -> str:
    if not base:
        base = "category"
    if base not in used_in_new and base not in EXISTING_SLUGS:
        used_in_new.add(base)
        return base
    if base in EXISTING_SLUGS:
        cand = f"{base}-saas"
        if len(cand) > 80:
            cand = cand[:80].rstrip("-")
        if cand not in used_in_new and cand not in EXISTING_SLUGS:
            used_in_new.add(cand)
            return cand
    for n in range(2, 1000):
        suffix = f"-{n}"
        cand = base if len(base) + len(suffix) <= 80 else base[: 80 - len(suffix)].rstrip("-")
        cand = f"{cand}{suffix}"
        if cand not in used_in_new and cand not in EXISTING_SLUGS:
            used_in_new.add(cand)
            return cand
    raise RuntimeError(f"could not allocate unique slug for {base!r}")


# Slug index: keyed by (level, name, *parent_names)
l1_slug = {}             # L1 name → final slug
l2_slug = {}             # (L1, L2) → final slug
l3_slug = {}             # (L1, L2, L3) → final slug

for name in l1_records:
    l1_slug[name] = pick_slug(slugify(name))
for l1, l2 in l2_records:
    l2_slug[(l1, l2)] = pick_slug(slugify(l2))
for l1, l2, l3 in l3_records:
    l3_slug[(l1, l2, l3)] = pick_slug(slugify(l3))


# ─── 5. Build emitted node list (with sort_order, parent_slug, db_level) ─────
emitted = []

for i, name in enumerate(l1_records, 1):
    emitted.append({
        "db_level": 2,
        "name": name,
        "slug": l1_slug[name],
        "parent_slug": SAAS_PARENT_SLUG,
        "parent_db_level": SAAS_PARENT_LEVEL,
        "sort_order": i * 10,
        "ancestors": [],
    })

# L2 sort_order: position within its L1 parent
l2_counter = {}
for l1, l2 in l2_records:
    l2_counter[l1] = l2_counter.get(l1, 0) + 10
    emitted.append({
        "db_level": 3,
        "name": l2,
        "slug": l2_slug[(l1, l2)],
        "parent_slug": l1_slug[l1],
        "parent_db_level": 2,
        "sort_order": l2_counter[l1],
        "ancestors": [{"db_level": 2, "name": l1}],
    })

# L3 sort_order: position within its (L1, L2) parent
l3_counter = {}
for l1, l2, l3 in l3_records:
    key = (l1, l2)
    l3_counter[key] = l3_counter.get(key, 0) + 10
    emitted.append({
        "db_level": 4,
        "name": l3,
        "slug": l3_slug[(l1, l2, l3)],
        "parent_slug": l2_slug[(l1, l2)],
        "parent_db_level": 3,
        "sort_order": l3_counter[key],
        "ancestors": [{"db_level": 2, "name": l1}, {"db_level": 3, "name": l2}],
    })

print(f"Total emitted nodes: {len(emitted)}")
by_db_level = {f"L{lvl}": sum(1 for n in emitted if n['db_level'] == lvl) for lvl in (2, 3, 4)}
print("By DB level:", by_db_level)


# ─── 6. Write JSON tree for listing remap step ───────────────────────────────
JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
JSON_OUT.write_text(
    json.dumps({"generated_from": EXCEL.name, "by_db_level": by_db_level, "nodes": emitted}, indent=2, ensure_ascii=False),
    encoding="utf-8",
)
print(f"Wrote {JSON_OUT}  ({len(emitted)} nodes)")


# ─── 7. Emit the SQL migration ───────────────────────────────────────────────
lines = []
lines.append("-- ============================================================")
lines.append("-- InfoWebWorld — Software & SaaS Taxonomy v2 Migration")
lines.append(f"-- Rebuilds the Software & SaaS sector with {len(emitted)} hierarchical")
lines.append("-- categories across 3 nested levels (DB L2..L4 under existing")
lines.append("-- 'software-saas' L1).")
lines.append("--")
lines.append("-- Source: Software-saas-taxonomy-full v1 final.xlsx")
lines.append("-- Run each section IN ORDER in phpMyAdmin.")
lines.append("-- ============================================================")
lines.append("")
lines.append("-- ═══ Section A: Safety ═════════════════════════════════════════")
lines.append("SET FOREIGN_KEY_CHECKS = 0;")
lines.append("")

lines.append("-- ═══ Section B: Disconnect existing software-saas submissions ═")
lines.append("-- Listings are PRESERVED. We null out category_id + listing_type_id")
lines.append("-- here, then re-attach them in Section F (appended at the bottom).")
lines.append("UPDATE submissions")
lines.append("   SET category_id = NULL, listing_type_id = NULL")
lines.append(" WHERE category_id IN (")
lines.append("   SELECT id FROM (")
lines.append("     SELECT c.id FROM categories c")
lines.append("      LEFT JOIN categories p   ON p.id   = c.parent_id")
lines.append("      LEFT JOIN categories gp  ON gp.id  = p.parent_id")
lines.append("      LEFT JOIN categories ggp ON ggp.id = gp.parent_id")
lines.append("      WHERE c.parent_id  = (SELECT id FROM categories WHERE slug='software-saas' AND level=1)")
lines.append("         OR p.parent_id  = (SELECT id FROM categories WHERE slug='software-saas' AND level=1)")
lines.append("         OR gp.parent_id = (SELECT id FROM categories WHERE slug='software-saas' AND level=1)")
lines.append("         OR ggp.parent_id = (SELECT id FROM categories WHERE slug='software-saas' AND level=1)")
lines.append("   ) AS saas_ids")
lines.append(" );")
lines.append("")

lines.append("-- ═══ Section C: Delete old software-saas dependents + categories ═══")
lines.append("-- Order: SEO content → listing_types → child cats → parent cats.")
lines.append("-- JOIN-based (no scalar subquery on same table) to dodge MySQL #1093.")
lines.append("")
lines.append("-- C.1: delete category_seo_content for old software-saas categories")
lines.append("DELETE sc FROM category_seo_content sc")
lines.append("  JOIN categories c ON c.id = sc.category_id")
lines.append("  LEFT JOIN categories p  ON p.id  = c.parent_id")
lines.append("  LEFT JOIN categories gp ON gp.id = p.parent_id")
lines.append(" WHERE c.parent_id  = (SELECT id FROM categories WHERE slug='software-saas' AND level=1)")
lines.append("    OR p.parent_id  = (SELECT id FROM categories WHERE slug='software-saas' AND level=1)")
lines.append("    OR gp.parent_id = (SELECT id FROM categories WHERE slug='software-saas' AND level=1);")
lines.append("")
lines.append("-- C.2: delete listing_types tied to old software-saas categories")
lines.append("DELETE lt FROM listing_types lt")
lines.append("  JOIN categories c ON c.id = lt.category_id")
lines.append("  LEFT JOIN categories p ON p.id = c.parent_id")
lines.append(" WHERE c.parent_id = (SELECT id FROM categories WHERE slug='software-saas' AND level=1)")
lines.append("    OR p.parent_id = (SELECT id FROM categories WHERE slug='software-saas' AND level=1);")
lines.append("")
lines.append("-- C.3: delete L3 categories under software-saas (JOIN-based)")
lines.append("DELETE c FROM categories c")
lines.append("  JOIN categories p  ON p.id  = c.parent_id")
lines.append("  JOIN categories gp ON gp.id = p.parent_id")
lines.append(" WHERE c.level = 3 AND gp.slug = 'software-saas' AND gp.level = 1;")
lines.append("")
lines.append("-- C.4: delete L2 categories under software-saas. JOIN-based.")
lines.append("DELETE c FROM categories c")
lines.append("  JOIN categories p ON p.id = c.parent_id")
lines.append(" WHERE c.level = 2 AND p.slug = 'software-saas' AND p.level = 1;")
lines.append("")

for db_lvl in (2, 3, 4):
    bucket = [n for n in emitted if n["db_level"] == db_lvl]
    lines.append(f"-- ═══ Section D.{db_lvl - 1}: Insert {len(bucket)} new L{db_lvl} categories ═══")
    for n in bucket:
        name_sql = sql_escape(n["name"])
        slug_sql = sql_escape(n["slug"])
        parent_slug_sql = sql_escape(n["parent_slug"])
        lines.append(
            "INSERT INTO categories (name, slug, level, parent_id, color, is_active, is_launched, is_navigation, sort_order)"
        )
        lines.append(
            f"SELECT '{name_sql}', '{slug_sql}', {db_lvl}, id, '{SAAS_COLOR}', 1, 1, 1, {n['sort_order']}"
        )
        lines.append(
            f"  FROM categories WHERE slug = '{parent_slug_sql}' AND level = {n['parent_db_level']} LIMIT 1;"
        )
    lines.append("")

lines.append("-- ═══ Section E: Re-enable FKs ═════════════════════════════════")
lines.append("SET FOREIGN_KEY_CHECKS = 1;")
lines.append("")
lines.append("-- ═══ Section F (appended by listing-remap step) ═══════════════")
lines.append("-- See bottom of file for UPDATE statements re-attaching live")
lines.append("-- software-saas submissions to their best-fit new categories.")
lines.append("")

SQL_OUT.parent.mkdir(parents=True, exist_ok=True)
SQL_OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {SQL_OUT}  ({SQL_OUT.stat().st_size:,} bytes)")
