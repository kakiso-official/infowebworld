"""
Build the Local Businesses v2 taxonomy migration SQL.

Excel: 3 cols (L1 Category | L2 Subcategory | L3 Leaf), 761 rows, clean.
Maps:
  Excel L1 → DB L2 (children of 'local-businesses' L1)
  Excel L2 → DB L3
  Excel L3 → DB L4

Run:  python scripts/build-local-businesses-v2-migration.py
"""
import json, re, sys, io
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path(r"C:\Users\AADIL PARMAR\Downloads\Local_Business_Structure v1.xlsx")
SQL_OUT = ROOT / "database" / "migration-local-businesses-taxonomy-v2.sql"
JSON_OUT = ROOT / "exports" / "local-businesses-taxonomy-v2.json"
NON_LB_SLUGS_FILE = ROOT / "exports" / "non-local-businesses-slugs.json"

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

LB_COLOR = "#F59E0B"             # amber
LB_PARENT_SLUG = "local-businesses"
LB_PARENT_LEVEL = 1


def slugify(name):
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    if len(s) > 80:
        s = s[:80].rstrip("-")
    return s


def sql_escape(s):
    return (s or "").replace("\\", "\\\\").replace("'", "''")


existing = json.loads(NON_LB_SLUGS_FILE.read_text(encoding="utf-8"))
EXISTING_SLUGS = {row["slug"] for row in existing["non_local_businesses_slugs"]}
print(f"Loaded {len(EXISTING_SLUGS)} existing non-local-businesses slugs.")

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["Master Taxonomy"]
print(f"Sheet: Master Taxonomy  rows={ws.max_row}")

raw = []
for r in range(2, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    if not (a and b and c):
        continue
    raw.append((str(a).strip(), str(b).strip(), str(c).strip()))
print(f"Valid rows: {len(raw)}")

seen_l1, seen_l2, seen_l3 = {}, {}, set()
l1_recs, l2_recs, l3_recs = [], [], []
for l1, l2, l3 in raw:
    if l1 not in seen_l1:
        seen_l1[l1] = True
        l1_recs.append(l1)
    if (l1, l2) not in seen_l2:
        seen_l2[(l1, l2)] = True
        l2_recs.append((l1, l2))
    triple = (l1, l2, l3)
    if triple not in seen_l3:
        seen_l3.add(triple)
        l3_recs.append(triple)
print(f"Unique L1: {len(l1_recs)}  L2: {len(l2_recs)}  L3: {len(l3_recs)}")

used = set()


def pick_slug(base):
    if not base:
        base = "category"
    if base not in used and base not in EXISTING_SLUGS:
        used.add(base)
        return base
    if base in EXISTING_SLUGS:
        cand = f"{base}-local"
        if len(cand) > 80:
            cand = cand[:80].rstrip("-")
        if cand not in used and cand not in EXISTING_SLUGS:
            used.add(cand)
            return cand
    for n in range(2, 1000):
        suffix = f"-{n}"
        cand = base if len(base) + len(suffix) <= 80 else base[: 80 - len(suffix)].rstrip("-")
        cand = f"{cand}{suffix}"
        if cand not in used and cand not in EXISTING_SLUGS:
            used.add(cand)
            return cand
    raise RuntimeError(f"could not allocate slug for {base!r}")


l1_slug = {n: pick_slug(slugify(n)) for n in l1_recs}
l2_slug = {(l1, l2): pick_slug(slugify(l2)) for l1, l2 in l2_recs}
l3_slug = {(l1, l2, l3): pick_slug(slugify(l3)) for l1, l2, l3 in l3_recs}

emitted = []
for i, n in enumerate(l1_recs, 1):
    emitted.append({"db_level": 2, "name": n, "slug": l1_slug[n], "parent_slug": LB_PARENT_SLUG, "parent_db_level": LB_PARENT_LEVEL, "sort_order": i * 10, "ancestors": []})
l2_counter = {}
for l1, l2 in l2_recs:
    l2_counter[l1] = l2_counter.get(l1, 0) + 10
    emitted.append({"db_level": 3, "name": l2, "slug": l2_slug[(l1, l2)], "parent_slug": l1_slug[l1], "parent_db_level": 2, "sort_order": l2_counter[l1], "ancestors": [{"db_level": 2, "name": l1}]})
l3_counter = {}
for l1, l2, l3 in l3_recs:
    k = (l1, l2)
    l3_counter[k] = l3_counter.get(k, 0) + 10
    emitted.append({"db_level": 4, "name": l3, "slug": l3_slug[(l1, l2, l3)], "parent_slug": l2_slug[(l1, l2)], "parent_db_level": 3, "sort_order": l3_counter[k], "ancestors": [{"db_level": 2, "name": l1}, {"db_level": 3, "name": l2}]})

print(f"Total emitted: {len(emitted)}")
by_db = {f"L{lvl}": sum(1 for x in emitted if x["db_level"] == lvl) for lvl in (2, 3, 4)}
print("By DB level:", by_db)

JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
JSON_OUT.write_text(json.dumps({"generated_from": EXCEL.name, "by_db_level": by_db, "nodes": emitted}, indent=2, ensure_ascii=False), encoding="utf-8")

lines = []
lines.append("-- ============================================================")
lines.append("-- InfoWebWorld — Local Businesses Taxonomy v2 Migration")
lines.append(f"-- Rebuilds the Local Businesses sector with {len(emitted)} hierarchical")
lines.append("-- categories across 3 nested levels (DB L2..L4 under existing")
lines.append("-- 'local-businesses' L1).")
lines.append("--")
lines.append("-- Source: Local_Business_Structure v1.xlsx")
lines.append("-- Run each section IN ORDER in phpMyAdmin.")
lines.append("-- ============================================================")
lines.append("")
lines.append("-- ═══ Section A: Safety ═════════════════════════════════════════")
lines.append("SET FOREIGN_KEY_CHECKS = 0;")
lines.append("")
lines.append("-- ═══ Section B: Disconnect existing local-businesses submissions ═══")
lines.append("UPDATE submissions")
lines.append("   SET category_id = NULL, listing_type_id = NULL")
lines.append(" WHERE category_id IN (")
lines.append("   SELECT id FROM (")
lines.append("     SELECT c.id FROM categories c")
lines.append("      LEFT JOIN categories p   ON p.id   = c.parent_id")
lines.append("      LEFT JOIN categories gp  ON gp.id  = p.parent_id")
lines.append("      LEFT JOIN categories ggp ON ggp.id = gp.parent_id")
lines.append("      WHERE c.parent_id  = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1)")
lines.append("         OR p.parent_id  = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1)")
lines.append("         OR gp.parent_id = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1)")
lines.append("         OR ggp.parent_id = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1)")
lines.append("   ) AS lb_ids")
lines.append(" );")
lines.append("")
lines.append("-- ═══ Section C: Delete old local-businesses dependents + categories ═══")
lines.append("")
lines.append("-- C.1: SEO content")
lines.append("DELETE sc FROM category_seo_content sc")
lines.append("  JOIN categories c ON c.id = sc.category_id")
lines.append("  LEFT JOIN categories p  ON p.id  = c.parent_id")
lines.append("  LEFT JOIN categories gp ON gp.id = p.parent_id")
lines.append(" WHERE c.parent_id  = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1)")
lines.append("    OR p.parent_id  = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1)")
lines.append("    OR gp.parent_id = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1);")
lines.append("")
lines.append("-- C.2: listing_types")
lines.append("DELETE lt FROM listing_types lt")
lines.append("  JOIN categories c ON c.id = lt.category_id")
lines.append("  LEFT JOIN categories p ON p.id = c.parent_id")
lines.append(" WHERE c.parent_id = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1)")
lines.append("    OR p.parent_id = (SELECT id FROM categories WHERE slug='local-businesses' AND level=1);")
lines.append("")
lines.append("-- C.3: L3 categories (JOIN-based)")
lines.append("DELETE c FROM categories c")
lines.append("  JOIN categories p  ON p.id  = c.parent_id")
lines.append("  JOIN categories gp ON gp.id = p.parent_id")
lines.append(" WHERE c.level = 3 AND gp.slug = 'local-businesses' AND gp.level = 1;")
lines.append("")
lines.append("-- C.4: L2 categories")
lines.append("DELETE c FROM categories c")
lines.append("  JOIN categories p ON p.id = c.parent_id")
lines.append(" WHERE c.level = 2 AND p.slug = 'local-businesses' AND p.level = 1;")
lines.append("")

for db_lvl in (2, 3, 4):
    bucket = [n for n in emitted if n["db_level"] == db_lvl]
    lines.append(f"-- ═══ Section D.{db_lvl - 1}: Insert {len(bucket)} new L{db_lvl} categories ═══")
    for n in bucket:
        name_sql = sql_escape(n["name"])
        slug_sql = sql_escape(n["slug"])
        parent_slug_sql = sql_escape(n["parent_slug"])
        lines.append("INSERT INTO categories (name, slug, level, parent_id, color, is_active, is_launched, is_navigation, sort_order)")
        lines.append(f"SELECT '{name_sql}', '{slug_sql}', {db_lvl}, id, '{LB_COLOR}', 1, 1, 1, {n['sort_order']}")
        lines.append(f"  FROM categories WHERE slug = '{parent_slug_sql}' AND level = {n['parent_db_level']} LIMIT 1;")
    lines.append("")

lines.append("-- ═══ Section E: Re-enable FKs ═════════════════════════════════")
lines.append("SET FOREIGN_KEY_CHECKS = 1;")
lines.append("")
lines.append("-- ═══ Section F: (no live listings to re-attach) ═══════════════")
lines.append("-- The inspect script found 0 submissions under local-businesses.")
lines.append("")

SQL_OUT.parent.mkdir(parents=True, exist_ok=True)
SQL_OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {SQL_OUT}  ({SQL_OUT.stat().st_size:,} bytes)")
